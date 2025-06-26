import logging
from urllib.parse import urlparse
import aiohttp
import pytz
from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, BufferedInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from aiohttp import ClientError, ClientConnectionError, ServerTimeoutError
import socket
from admin.keyboards.admin_reply import admin_keyboard
import pandas as pd
from io import BytesIO
from datetime import datetime as dt
from client.keyboards.reply import main_kb
from data.url import url_users
from data.config import config_settings
from utils.filters import ChatTypeFilter, IsGroupAdmin, ADMIN_CHAT_ID

logger = logging.getLogger(__name__)


admin_router = Router()
admin_router.message.filter(
    ChatTypeFilter("private"),
    IsGroupAdmin([ADMIN_CHAT_ID], show_message=True)
)


async def generate_excel_report():
    """Генерирует отчёт в формате Excel с данными пользователей.

    Returns:
        BytesIO: Объект с Excel-файлом или None, если данные отсутствуют.

    Raises:
        Exception: Если произошла ошибка при запросе API или создании файла.
    """
    if not config_settings.BOT_API_KEY:
        logger.error("BOT_API_KEY не установлен")
        return None

    async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
        # Получаем всех пользователей
        try:
            async with session.get(
                url_users,
                headers={"X-Bot-Api-Key": config_settings.BOT_API_KEY.get_secret_value()}
            ) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    logger.error(f"Ошибка при получении пользователей: {resp.status}, {error_text}")
                    return None
                users = await resp.json()
                logger.info(f"Получено {len(users)} пользователей")
        except Exception as e:
            logger.exception(f"Ошибка запроса пользователей: {e}")
            return None

    if not users:
        logger.info("Нет данных пользователей для отчета")
        return None

    # Создаем DataFrame
    df = pd.DataFrame(users)

    # Исключаем ненужные столбцы
    columns_to_drop = ['id', 'password', 'groups', 'user_permissions']
    df = df.drop(columns=columns_to_drop, errors='ignore')

    # Применяем преобразования к полям
    bool_columns = ['is_bot', 'is_staff', 'is_active', 'is_superuser']
    for col in bool_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: 'Да' if x else 'Нет')

    datetime_columns = ['date_joined', 'last_activity']
    for col in datetime_columns:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: x[:19].replace('T', ' ') if isinstance(x, str) else None
            )

    # Форматируем birth_date
    if 'birth_date' in df.columns:
        df['birth_date'] = df['birth_date'].apply(
            lambda x: x if isinstance(x, str) and x else None
        )

    # Переименовываем колонки для красивого отображения
    column_mapping = {
        'tg_id': 'TG ID',
        'username': 'Никнейм',
        'first_name': 'Имя (Telegram)',
        'last_name': 'Фамилия (Telegram)',
        'user_first_name': 'Имя (карта)',
        'user_last_name': 'Фамилия (карта)',
        'birth_date': 'Дата рождения',
        'email': 'Email',
        'phone_number': 'Номер телефона',
        'is_bot': 'Бот',
        'date_joined': 'Дата регистрации',
        'last_activity': 'Последняя активность',
        'is_staff': 'Админ',
        'is_active': 'Активный',
        'is_superuser': 'Суперюзер',
        'role': 'Роль'
    }
    df.rename(columns=column_mapping, inplace=True)

    # Ограничиваем колонки, чтобы включить только нужные
    columns_to_keep = [
        'TG ID', 'Никнейм', 'Имя (Telegram)', 'Фамилия (Telegram)',
        'Имя (карта)', 'Фамилия (карта)', 'Дата рождения', 'Email', 'Номер телефона',
        'Роль', 'Активный'
    ]
    df = df[[col for col in columns_to_keep if col in df.columns]]

    # Создаем Excel-файл
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Пользователи')
        worksheet = writer.sheets['Пользователи']

        # Настройка выравнивания
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment

        # Автоподбор ширины колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Стилизация заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    output.seek(0)
    logger.info("Excel-отчет по всем пользователям успешно сгенерирован")
    return output

@admin_router.message(Command("admin"))
async def admin_panel(message: Message, bot: Bot):
    """Открывает админ-панель для пользователя.

    Args:
        message (Message): Сообщение от пользователя с командой /admin.
        bot (Bot): Объект бота для отправки ответа.

    Notes:
        Доступно только администраторам через фильтры.
    """
    user_id = message.from_user.id
    await message.answer(
        "🛠 Админ-панель",
        reply_markup=admin_keyboard()
    )

@admin_router.message(F.text == "📊 Статистика")
async def show_statistics(message: Message):
    """Отображает статистику пользователей и предлагает выгрузку в Excel.

    Args:
        message (Message): Сообщение от пользователя с запросом статистики.

    Notes:
        Выполняет запрос к API и строит инлайн-клавиатуру с кнопкой выгрузки.
    """
    try:
        # Проверка соединения с сервером
        try:
            parsed_url = urlparse(url_users)
            host = parsed_url.hostname
            port = parsed_url.port or (80 if parsed_url.scheme == 'http' else 443)
            with socket.create_connection((host, port), timeout=3):
                pass
        except (socket.timeout, ConnectionRefusedError, OSError) as e:
            logger.error(f"Сервер недоступен: {e}")
            await message.answer("🔴 Сервер статистики временно недоступен. Попробуйте позже.")
            return
        except Exception as e:
            logger.error(f"Ошибка проверки соединения: {e}")
            await message.answer("⚠️ Ошибка при проверке доступности сервера")
            return

        if not config_settings.BOT_API_KEY:
            logger.error("BOT_API_KEY не установлен")
            await message.answer("⚠️ Ошибка конфигурации сервера")
            return

        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
            try:
                async with session.get(
                    url_users,
                    headers={"X-Bot-Api-Key": config_settings.BOT_API_KEY.get_secret_value()}
                ) as resp:
                    if resp.status != 200:
                        error_text = await resp.text()
                        logger.error(f"API error {resp.status}: {error_text}")
                        await message.answer(f"⚠️ Ошибка сервера {resp.status}. Попробуйте позже.")
                        return
                    try:
                        users = await resp.json()
                    except ValueError as e:
                        logger.error(f"Invalid JSON response: {e}")
                        await message.answer("⚠️ Ошибка обработки данных сервера")
                        return
            except ServerTimeoutError as e:
                logger.error(f"Timeout error: {e}")
                await message.answer("⏳ Сервер не отвечает. Попробуйте позже.")
                return
            except ClientConnectionError as e:
                logger.error(f"Connection error: {e}")
                await message.answer("🔴 Не удалось подключиться к серверу")
                return
            except ClientError as e:
                logger.error(f"Client error: {e}")
                await message.answer("⚠️ Ошибка при запросе к серверу")
                return

        # Обработка данных
        total_users = len(users)
        active_users = sum(1 for user in users if user.get('is_active', False))

        builder = InlineKeyboardBuilder()
        builder.add(InlineKeyboardButton(
            text="📥 Выгрузить в Excel",
            callback_data="export_users_excel"
        ))

        await message.answer(
            f"📊 <b>Статистика бота</b>\n\n"
            f"👥 Всего пользователей: <code>{total_users}</code>\n"
            f"🟢 Активных: <code>{active_users}</code>",
            reply_markup=builder.as_markup()
        )

    except Exception as e:
        logger.error(f"Unexpected error in show_statistics: {e}", exc_info=True)
        await message.answer("⚠️ Непредвиденная ошибка при получении статистики")

@admin_router.callback_query(F.data == "export_users_excel")
async def export_users_excel(callback: CallbackQuery):
    """Генерирует и отправляет отчёт пользователей в формате Excel.

    Args:
        callback (CallbackQuery): Callback-запрос от кнопки "Выгрузить в Excel".

    Notes:
        Использует generate_excel_report для создания файла.
    """
    try:
        await callback.answer("⏳ Готовим отчет...")
        excel_file = await generate_excel_report()

        if excel_file:
            await callback.message.answer_document(
                BufferedInputFile(
                    excel_file.getvalue(),
                    filename=f"users_report_{dt.now(tz=pytz.timezone('Europe/Moscow')).strftime('%Y%m%d_%H%M%S')}.xlsx"
                ),
                caption="📊 Отчет по всем пользователям"
            )
        else:
            await callback.message.answer("❌ Нет данных для выгрузки")
    except Exception as e:
        logger.error(f"Error in export_users_excel: {e}", exc_info=True)
        await callback.message.answer("⚠️ Произошла ошибка при генерации отчета")
    finally:
        await callback.answer()


@admin_router.message(F.text == "Выход")
async def exit_admin_panel(message: Message):
    """Выходит из админ-панели и возвращает основное меню.

    Args:
        message (Message): Сообщение от пользователя с запросом выхода.

    Notes:
        Устанавливает основную клавиатуру main_kb.
    """
    await message.answer(
        "Выход из админ-панели",
        reply_markup=main_kb
    )
